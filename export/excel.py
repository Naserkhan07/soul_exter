"""
Excel exporter — SQLite is the database, Excel is the deliverable.

    python main.py export   ->  output/india_leads.xlsx
"""

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from database.database import LeadDatabase

log = logging.getLogger("export.excel")

COLUMNS = [
    ("Lead ID", "lead_id", 16),
    ("Business Name", "business_name", 28),
    ("Business Category", "business_category", 20),
    ("Person Name", "person_name", 18),
    ("Person Role", "person_role", 16),
    ("State", "state", 16),
    ("District", "district", 14),
    ("City", "city", 14),
    ("Locality", "locality", 14),
    ("Full Location", "full_location", 32),
    ("Phone", "phone", 16),
    ("WhatsApp", "whatsapp", 16),
    ("Email", "email", 26),
    ("Website", "website", 30),
    ("LinkedIn", "linkedin", 30),
    ("Instagram", "instagram", 26),
    ("Facebook", "facebook", 26),
    ("YouTube", "youtube", 26),
    ("Twitter/X", "twitter", 22),
    ("Google Maps", "google_maps", 30),
    ("Other Contact URL", "other_contact_url", 24),
    ("Source URLs", "source_urls", 40),
    ("Digital Marketing Need", "digital_marketing_need", 12),
    ("SEO Need", "seo_need", 10),
    ("Local SEO Need", "local_seo_need", 10),
    ("Social Media Need", "social_media_need", 12),
    ("Website Need", "website_need", 10),
    ("E-commerce Need", "ecommerce_need", 10),
    ("Mobile App Need", "mobile_app_need", 10),
    ("Web App Need", "web_app_need", 10),
    ("AI Automation Need", "ai_automation_need", 10),
    ("Tech Support Need", "technical_support_need", 10),
    ("Detected Problems", "detected_problems", 45),
    ("Recommended Services", "recommended_services", 35),
    ("Lead Score", "lead_score", 10),
    ("Evidence / Reason", "evidence_reason", 50),
    ("Discovery Source", "discovery_source", 14),
    ("Date Found", "date_found", 12),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SCORE_FILLS = {
    "high": PatternFill("solid", fgColor="C6EFCE"),    # >= 70
    "medium": PatternFill("solid", fgColor="FFEB9C"),  # 40-69
}


def export_to_excel(min_score: int = 0, path=None) -> str:
    db = LeadDatabase()
    leads = db.all_leads(min_score=min_score)
    path = str(path or config.EXCEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "India Leads"
    ws.freeze_panes = "C2"

    for col, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_i, lead in enumerate(leads, start=2):
        for col, (_, attr, _w) in enumerate(COLUMNS, start=1):
            value = getattr(lead, attr, "")
            if isinstance(value, list):
                value = "; ".join(str(v) for v in value)
            cell = ws.cell(row=row_i, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(
                attr in ("detected_problems", "evidence_reason",
                         "source_urls", "recommended_services")))
        score = lead.lead_score or 0
        score_col = next(i for i, c in enumerate(COLUMNS, 1)
                         if c[1] == "lead_score")
        if score >= 70:
            ws.cell(row=row_i, column=score_col).fill = _SCORE_FILLS["high"]
        elif score >= 40:
            ws.cell(row=row_i, column=score_col).fill = _SCORE_FILLS["medium"]

    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    log.info("Exported %d leads to %s", len(leads), path)
    return path
